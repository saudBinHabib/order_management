import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

BASE_DIR = Path(__file__).parent
BESTELLLISTE_DIR = Path(os.environ.get("BESTELLLISTE_DIR", BASE_DIR / "bestellliste"))
ORDERS_DIR = Path(os.environ.get("ORDERS_DIR", BASE_DIR / "orders"))
ORDERS_DIR.mkdir(exist_ok=True)

TEMPLATE_FILES = {
    "Feinfood Bestellliste": BESTELLLISTE_DIR / "fein food .xlsx",
    "Nachoking Bestellliste": BESTELLLISTE_DIR / "nacho king.xlsx",
}

NAME_COL = 2       # column B: item name (TK-Ware / Naschokings)
BESTELLUNG_COL = 5  # column E: Bestellung (order quantity), same in both templates

SUPPLIER = {
    "name": "FeinFood Express GmbH",
    "address": "Flinschstraße 2-4, 60388 Frankfurt am Main",
    "phone": "+49 176 317 22918",
    "email": "info@main-food.com",
    "bank": "Commerzbank FFM GF-F48",
    "iban": "DE51500400480170017800",
    "bic": "COBADEFFXXX",
    "ustid": "DE361848916",
}

CUSTOMER = {
    "name": "Habib RFC Mannheim GmbH",
    "contact": "Daud Bin Habib Khan",
    "address_line1": "Q 4 6-8",
    "address_line2": "68161 Mannheim",
    "kunden_nr": "10016",
}


def _slug(bestellliste: str) -> str:
    return "feinfood" if bestellliste.startswith("Feinfood") else "nachoking"


def generate_vendor_order_file(bestellliste: str, items: list[dict], order_date: str, order_id: int) -> Path:
    """Fill quantities into the vendor's own bestellliste template, leaving all
    other rows untouched. `items` = list of {matched_item_name, quantity}."""
    template_path = TEMPLATE_FILES[bestellliste]
    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]

    name_to_qty = {it["matched_item_name"]: it["quantity"] for it in items}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell_name = row[NAME_COL - 1].value
        if cell_name is None:
            continue
        cell_name = str(cell_name).strip()
        if cell_name in name_to_qty:
            ws.cell(row=row[0].row, column=BESTELLUNG_COL, value=name_to_qty[cell_name])

    safe_date = order_date.replace("-", "")
    out_path = ORDERS_DIR / f"{_slug(bestellliste)}_vendor_order_{safe_date}_{order_id}.xlsx"
    wb.save(out_path)
    return out_path


def generate_invoice_pattern_file(bestellliste: str, items: list[dict], order_date: str, order_id: int) -> Path:
    """Build an order document that mirrors the FeinFood invoice layout:
    same header blocks, same table columns, same Netto/USt/Brutto footer.
    `items` = list of {product_number, product_name, unit, quantity, unit_price, vat_percent}.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bestellung"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center")

    ws.append([SUPPLIER["name"]])
    ws["A1"].font = title_font
    ws.append([SUPPLIER["address"]])
    ws.append([f"Tel. {SUPPLIER['phone']}"])
    ws.append([f"E-Mail: {SUPPLIER['email']}"])
    ws.append([])
    ws.append([CUSTOMER["name"]])
    ws.append([CUSTOMER["contact"]])
    ws.append([CUSTOMER["address_line1"]])
    ws.append([CUSTOMER["address_line2"]])
    ws.append([])
    ws.append(["Bestellung Nr.:", order_id])
    ws.append(["Kunden Nr.:", CUSTOMER["kunden_nr"]])
    ws.append(["Bestelldatum:", order_date])
    ws.append(["Bestellliste:", bestellliste])
    ws.append([])

    header_row = ws.max_row + 1
    headers = ["Pos", "Menge", "Einheit", "Nummer", "Text", "Einzelpreis EUR", "USt. %", "Gesamtpreis EUR"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = bold
        c.border = border
        c.alignment = center

    netto_total = 0.0
    vat_by_rate = {}
    for i, it in enumerate(items, start=1):
        line_netto = round(it["quantity"] * it["unit_price"], 2)
        netto_total += line_netto
        vat_by_rate.setdefault(it["vat_percent"], 0.0)
        vat_by_rate[it["vat_percent"]] += line_netto
        row = [
            i,
            it["quantity"],
            it["unit"] or "",
            it["product_number"],
            it["product_name"],
            it["unit_price"],
            it["vat_percent"],
            line_netto,
        ]
        ws.append(row)
        r = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = border

    ws.append([])
    netto_total = round(netto_total, 2)
    ws.append(["", "", "", "", "", "", "Gesamt Netto", netto_total])
    ws.cell(row=ws.max_row, column=7).font = bold

    vat_total = 0.0
    for rate in sorted(vat_by_rate):
        base = round(vat_by_rate[rate], 2)
        vat_amount = round(base * rate / 100, 2)
        vat_total += vat_amount
        ws.append(["", "", "", "", "", "", f"zzgl. {rate:.2f} % USt. auf {base:.2f}", vat_amount])

    vat_total = round(vat_total, 2)
    brutto_total = round(netto_total + vat_total, 2)
    ws.append([])
    ws.append(["", "", "", "", "", "", "Geschätzter Bestellwert (Brutto)", brutto_total])
    ws.cell(row=ws.max_row, column=7).font = bold
    ws.cell(row=ws.max_row, column=8).font = bold

    for col_letter, width in zip("ABCDEFGH", [6, 8, 10, 8, 40, 14, 10, 16]):
        ws.column_dimensions[col_letter].width = width

    safe_date = order_date.replace("-", "")
    out_path = ORDERS_DIR / f"{_slug(bestellliste)}_order_confirmation_{safe_date}_{order_id}.xlsx"
    wb.save(out_path)
    return out_path


def compute_totals(items: list[dict]):
    """items: list of {quantity, unit_price, vat_percent}. Returns (netto, vat, brutto)."""
    netto_total = 0.0
    vat_total = 0.0
    for it in items:
        line_netto = it["quantity"] * it["unit_price"]
        netto_total += line_netto
        vat_total += line_netto * it["vat_percent"] / 100
    netto_total = round(netto_total, 2)
    vat_total = round(vat_total, 2)
    brutto_total = round(netto_total + vat_total, 2)
    return netto_total, vat_total, brutto_total
